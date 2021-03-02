# coding: UTF-8
import datetime
from math import log, pow
from calculate_ad_data import calculate_V, calculate_bat_I, calculate_R, calculate_T, calculate_solar_I_minus_y_dep, calculate_solar_I_plus_y_undep, calculate_solar_I_plus_x, calculate_solar_I_minus_x, calculate_solar_I_plus_y_dep, calculate_solar_I_minus_z, calculate_txCW_FM_I, calculate_rx_modem_I, calculate_mis5V_I, calculate_mobc5V_I, calculate_misbus_I
import sys
import openpyxl
import os
import data_format
import decode
import wx
import time
import subprocess

def make_log(data, path):
    #decoded_time = str(datetime.datetime.now())
    if not os.path.exists('mewton_log'):
        os.mkdir('mewton_log')
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    index = ['テレメトリ', 'テレメトリID', 'ミッションモードID', 'BOBC時刻 [sec]', '異常ステータス', 'バッテリー電圧 [V]', 'バッテリー電流 [A]', 'バッテリー温度 [deg]', '-Y Deployed Solar Current [A]', '＋Y Undeployed Solar Current [A]', '＋X Solar Current [A]', '-X Solar Current [A]', '＋Y Deployed Solar [A]', '-Z Solar Current [A]', 'Tranceiver Temp [A]', 'Mission-Board Temp [deg]', '＋X Panel Temp [deg]', '＋Z Panel Temp [deg]', '＋Y Panel Temp [deg]', '-Y Panel Temp [deg]', '-X Panel Temp [deg]', '-Z Panel Temp [deg]', 'EPS1-Shunt Temp [deg]', 'TX(CW/FM) Current [A]', 'RX, Modem(RX) etc. Current [A]', 'Mission-Board 5V-Line Current [A]', 'MOBC 5V-Line Current [A]', 'Mission-Board Bus-Line Current [A]', 'Bus-System 5V-Line2 Voltage [V]', 'Bus-System Bus-Line Voltage [V]', 'RSSI [V]']
    for i in range(len(index)):
        ws.cell(1, i+1).value = index[i]

    for i in range(len(data)):
        for k in range(len(data[0])):
            ws.cell(i+2, k+1).value = data[i][k]

    filename = path.split('/')[-1]
    filename = filename.split('.txt')[0]
    filename = filename.split('.log')[0]
    filename = "./mewton_log/" + filename + ".xlsx"
    wb.save(filename)
    return filename

class FileDropTarget(wx.FileDropTarget):
    """ Drag & Drop Class """
    def __init__(self, window):
        wx.FileDropTarget.__init__(self)
        self.window = window

    def OnDropFiles(self, x, y, files):
        #self.window.text_entry.SetLabel(str(files))
        for file in files:
            #path = input()
            #path = 'BUS_function_AO40_GMSK9k6_2.txt'
            path = file
            f = open(path, 'r')
            if f.readline()[0] == '0':
                telems = data_format.nyanpath(path)
            else:
                telems = data_format.direwolf(path)
            filename = make_log(decode.decode_FP(telems), path)
        log_dir = os.getcwd() + '/mewton_log'
        if os.name == 'nt':
            subprocess.run(["explorer", log_dir])
        else:
            subprocess.run(["open", log_dir])
            #subprocess.run('explorer {}'.format(os.getcwd()))
        #time.sleep(1)
        #self.window.label.SetBackgroundColour("#ffc0cb")
        return 0

class App(wx.Frame):
    """ GUI """
    def __init__(self, parent, id, title):
        wx.Frame.__init__(self, parent, id, title, size=(500, 300), style=wx.DEFAULT_FRAME_STYLE)

        # パネル
        self.p = wx.Panel(self, wx.ID_ANY)

        self.label = wx.StaticText(self.p, wx.ID_ANY, '\n\n\nMEWTON\n[NYANPATH and Direwolf Log Decoader]\n\n\n\nDrop log file(s) here (.log or .txt)\n', style=wx.SIMPLE_BORDER | wx.TE_CENTER)
        self.label.SetBackgroundColour("#ffc0cb")
        font = wx.Font(15, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL)
        self.label.SetFont(font)
        # ドロップ対象の設定
        self.label.SetDropTarget(FileDropTarget(self))

        # テキスト入力ウィジット
        #self.text_entry = wx.TextCtrl(self.p, wx.ID_ANY)

        # レイアウト
        layout = wx.BoxSizer(wx.VERTICAL)
        layout.Add(self.label, flag=wx.EXPAND | wx.ALL, border=10, proportion=1)
        #layout.Add(self.text_entry, flag=wx.EXPAND | wx.ALL, border=10)
        self.p.SetSizer(layout)

        self.Show()

if __name__ == '__main__':
    app = wx.App()
    App(None, -1, 'MEWTON')
    app.MainLoop()
